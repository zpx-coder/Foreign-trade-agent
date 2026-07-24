"""客户管理 API — Phase 4 Customer Acquisition"""

import asyncio
import io
import json
import logging
import uuid as _uuid
from datetime import datetime
from ipaddress import ip_address, ip_network
from typing import Optional, List
from urllib.parse import urlparse

import httpx
from bs4 import BeautifulSoup
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import select, func, and_, update, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db, get_session, AsyncSessionLocal
from app.config import settings
from app.core.auth import get_current_user
from app.models.user import User
from app.models.icp import Icp
from app.models.customer import Customer
from app.models.contact import Contact
from app.schemas.customer import (
    ContactCreateRequest,
    ContactUpdateRequest,
    ContactResponse,
    CustomerCreateRequest,
    CustomerUpdateRequest,
    CustomerListItem,
    CustomerResponse,
    CustomerDetailResponse,
    CustomerListResponse,
    CustomerSearchRequest,
    BatchStatusRequest,
    CustomerImportRowError,
    CustomerImportResponse,
)
from app.services.ai_service import AIServiceError, get_ai_service
from app.services.ai.customer_extractor import CustomerExtractor
from app.services.enrichment.customer_enricher import CustomerEnricher
from app.services.enrichment.contact_scraper import ContactScraper
from app.services.enrichment.hunter_service import HunterService
from app.services.search.base import SearchResult
from app.services.search.google_channel import GoogleSearchChannel
from app.services.search.linkedin_channel import LinkedInSearchChannel
from app.services.search.linkedin_people_channel import LinkedInPeopleSearchChannel
from app.services.search.duckduckgo_channel import DuckDuckGoSearchChannel
from app.services.search.ai_search_channel import AISearchChannel
from app.services.search.serper_channel import SerperSearchChannel
from app.services.search.contact_search import ContactSearchService
from app.services.search.aggregator import SearchAggregator

logger = logging.getLogger(__name__)

router = APIRouter()

# ── 搜索限流常量 ──
_MAX_SEARCH_RESULTS = 20
_MAX_EXPORT_ROWS = 5000
_MAX_PAGE_FETCH_TIMEOUT = 10.0

# ── 私有 IP 段（SSRF 防护） ──
_PRIVATE_NETWORKS = [
    ip_network("10.0.0.0/8"),
    ip_network("172.16.0.0/12"),
    ip_network("192.168.0.0/16"),
    ip_network("127.0.0.0/8"),
    ip_network("169.254.0.0/16"),
    ip_network("0.0.0.0/8"),
]


# ── 辅助函数 ──

def _parse_uuid(val: str, label: str = "ID") -> _uuid.UUID:
    """将路径参数解析为 UUID，无效时返回 400"""
    try:
        return _uuid.UUID(val)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail=f"{label}格式无效")


def _is_safe_url(url: str) -> bool:
    """SSRF 防护：检查 URL 是否安全可访问"""
    if not url:
        return False
    try:
        parsed = urlparse(url if "://" in url else f"https://{url}")
        hostname = parsed.hostname
        if not hostname:
            return False
        # 仅允许 HTTP/HTTPS
        if parsed.scheme not in ("http", "https"):
            return False
        # 如果 hostname 本身就是 IP，直接校验
        try:
            addr = ip_address(hostname)
        except ValueError:
            # hostname 是域名，需要 DNS 解析后校验
            import socket
            try:
                addr = ip_address(socket.getaddrinfo(hostname, None)[0][4][0])
            except (ValueError, OSError, IndexError):
                # DNS 解析失败（可能内网域名），为安全起见放行
                # 后续 httpx 请求会自然失败
                return True
        for net in _PRIVATE_NETWORKS:
            if addr in net:
                return False
        return True
    except (ValueError, OSError):
        return False


def _build_customer_filter(
    tenant_id: _uuid.UUID,
    status: Optional[str] = None,
    source: Optional[str] = None,
    search: Optional[str] = None,
    icp_id: Optional[str] = None,
):
    """构建客户筛选条件（list + export 共用）"""
    conditions = [Customer.tenant_id == tenant_id]
    if status:
        conditions.append(Customer.status == status)
    if source:
        conditions.append(Customer.source == source)
    if search:
        conditions.append(
            or_(Customer.name.ilike(f"%{search}%"), Customer.industry.ilike(f"%{search}%"))
        )
    if icp_id:
        if icp_id == "__none__":
            conditions.append(Customer.icp_id.is_(None))
        else:
            conditions.append(Customer.icp_id == icp_id)
    where_clause = conditions[0]
    for c in conditions[1:]:
        where_clause = where_clause & c
    return where_clause

def _row_to_listitem(customer: Customer) -> CustomerListItem:
    """将 Customer ORM 对象转为列表项（含联系人计数 + ICP 名称）"""
    icp_name = None
    if hasattr(customer, "icp") and customer.icp:
        icp_name = customer.icp.name
    contacts = customer.contacts or []
    contacts_with_email = sum(
        1 for c in contacts if c.email and c.email.strip()
    )
    return CustomerListItem(
        id=customer.id,
        name=customer.name,
        industry=customer.industry,
        country=customer.country,
        source=customer.source,
        status=customer.status,
        website=customer.website,
        contacts_count=len(contacts),
        contacts_with_email_count=contacts_with_email,
        icp_id=customer.icp_id,
        icp_name=icp_name,
        created_at=customer.created_at,
        updated_at=customer.updated_at,
    )


def _row_to_response(customer: Customer) -> CustomerResponse:
    """将 Customer ORM 对象转为响应"""
    return CustomerResponse(
        id=customer.id,
        tenant_id=customer.tenant_id,
        name=customer.name,
        industry=customer.industry,
        website=customer.website,
        country=customer.country,
        city=customer.city,
        company_size=customer.company_size,
        description=customer.description,
        source=customer.source,
        source_url=customer.source_url,
        icp_id=customer.icp_id,
        status=customer.status,
        source_data=customer.source_data,
        ai_summary=customer.ai_summary,
        notes=customer.notes,
        contacts_count=len(customer.contacts) if customer.contacts else 0,
        created_at=customer.created_at,
        updated_at=customer.updated_at,
    )


# ── 客户 CRUD ──

@router.get("", response_model=CustomerListResponse)
async def list_customers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=2000),
    status: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    search: Optional[str] = Query(None, description="搜索公司名/行业"),
    icp_id: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """客户列表（分页 + 筛选 + 搜索）"""
    where_clause = _build_customer_filter(
        current_user.tenant_id, status, source, search, icp_id
    )

    # 总数
    count_q = select(func.count(Customer.id)).where(where_clause)
    total = (await db.execute(count_q)).scalar() or 0

    # 分页
    offset = (page - 1) * page_size
    q = (
        select(Customer)
        .where(where_clause)
        .options(selectinload(Customer.contacts), selectinload(Customer.icp))
        .order_by(Customer.created_at.desc())
        .offset(offset)
        .limit(page_size)
    )
    rows = (await db.execute(q)).scalars().all()

    return CustomerListResponse(
        items=[_row_to_listitem(r) for r in rows],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.post("", response_model=CustomerDetailResponse, status_code=status.HTTP_201_CREATED)
async def create_customer(
    data: CustomerCreateRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """手动创建客户（含可选联系人列表）"""
    payload = data.model_dump(exclude={"contacts"})
    contacts_payload = data.contacts or []

    customer = Customer(
        tenant_id=current_user.tenant_id,
        created_by=current_user.id,
        **payload,
    )
    db.add(customer)
    await db.flush()  # 获取 customer.id

    for c_data in contacts_payload:
        contact = Contact(
            customer_id=customer.id,
            tenant_id=current_user.tenant_id,
            **c_data.model_dump(),
        )
        db.add(contact)

    await db.commit()
    await db.refresh(customer)
    # eager-load contacts
    await db.refresh(customer, ["contacts"])

    return CustomerDetailResponse(
        **_row_to_response(customer).model_dump(),
        contacts=[ContactResponse.model_validate(c) for c in customer.contacts],
    )


# ── Excel 导入 ──（注意：路由注册须在 /{customer_id} 之前，避免路径被参数捕获）

# 导入用列名映射（中文列名 → Customer 字段）
_IMPORT_COLUMN_MAP = {
    "公司名称": "name",
    "行业": "industry",
    "网站": "website",
    "官网": "website",
    "国家": "country",
    "城市": "city",
    "规模": "company_size",
    "公司规模": "company_size",
    "描述": "description",
    "公司描述": "description",
    "备注": "notes",
    "来源url": "source_url",
    "来源网址": "source_url",
    "状态": "status",
}

_IMPORT_REQUIRED_COLUMNS = {"公司名称"}


@router.get("/import-template")
async def download_import_template():
    """下载 Excel 导入模板（无需登录）"""
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "客户导入模板"

    # 标题行样式
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        ("公司名称", 22, True),
        ("行业", 14),
        ("网站", 22),
        ("国家", 10),
        ("城市", 12),
        ("规模", 10),
        ("描述", 30),
        ("备注", 20),
        ("来源URL", 25),
        ("状态", 10),
    ]

    for col_idx, (title, width, *_) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 28

    # 示例行
    example_data = [
        "示例科技有限公司", "机械制造", "https://www.example.com",
        "德国", "慕尼黑", "50-200", "工业自动化设备制造商", "展会认识",
        "https://www.linkedin.com/company/example", "new",
    ]
    for col_idx, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_idx, value=value)

    # 状态数据验证
    from openpyxl.worksheet.datavalidation import DataValidation
    status_dv = DataValidation(
        type="list",
        formula1='"new,contacted,qualified,negotiating,closed,rejected"',
        allow_blank=True,
    )
    status_dv.prompt = "请选择状态"
    status_dv.promptTitle = "客户状态"
    ws.add_data_validation(status_dv)
    status_dv.add(ws.cell(row=2, column=10))
    # 将验证应用到后续行
    for r in range(3, 102):
        status_dv.add(ws.cell(row=r, column=10))

    ws.auto_filter.ref = f"A1:J1"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_import_template.xlsx"},
    )


@router.post("/import", response_model=CustomerImportResponse)
async def import_customers(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """通过 Excel 批量导入客户"""
    # 校验文件格式
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式的 Excel 文件")

    contents = await file.read()
    if len(contents) == 0:
        raise HTTPException(status_code=400, detail="文件为空")

    # 文件大小限制 5MB
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过 5MB")

    try:
        wb = Workbook.__new__(Workbook)
        # 使用 load_workbook 读取上传文件
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="工作簿中没有工作表")
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件，请确认文件格式正确")

    # 读取表头
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel 文件为空")

    if not header_row:
        raise HTTPException(status_code=400, detail="无法读取表头行")

    # 规范化表头并建立列索引映射
    col_index: dict[int, str] = {}
    for idx, col_name in enumerate(header_row):
        if col_name and isinstance(col_name, str):
            normalized = col_name.strip()
            field = _IMPORT_COLUMN_MAP.get(normalized, normalized.lower())
            col_index[idx] = field

    # 检查必要列
    has_name_column = any(
        field == "name" for field in col_index.values()
    )
    if not has_name_column:
        raise HTTPException(
            status_code=400,
            detail="Excel 缺少「公司名称」列，请使用模板文件或确保第一列是公司名称",
        )

    created = 0
    skipped = 0
    errors: list[CustomerImportRowError] = []
    max_rows = 1000

    for row_num, row_data in enumerate(rows_iter, 2):  # 从第 2 行开始（1-indexed）
        if row_num > max_rows + 1:
            break

        if not row_data or all(v is None or str(v).strip() == "" for v in row_data):
            continue  # 跳过空行

        # 构建客户数据
        customer_data: dict = {"source": "manual_import", "status": "new"}
        for col_idx, field_name in col_index.items():
            if col_idx < len(row_data) and row_data[col_idx] is not None:
                value = str(row_data[col_idx]).strip()
                if value and value.lower() not in ("none", "null", "n/a", "-"):
                    customer_data[field_name] = value

        name = customer_data.pop("name", "").strip()
        if not name:
            errors.append(CustomerImportRowError(row=row_num, message="公司名称为空，已跳过"))
            skipped += 1
            continue

        # 校验状态值
        valid_statuses = {"new", "contacted", "qualified", "negotiating", "closed", "rejected"}
        status_val = customer_data.get("status", "new")
        if status_val not in valid_statuses:
            customer_data["status"] = "new"

        # 去重：按公司名 + 租户检查
        existing = await db.execute(
            select(func.count(Customer.id)).where(
                Customer.tenant_id == current_user.tenant_id,
                Customer.name == name,
            )
        )
        if (existing.scalar() or 0) > 0:
            skipped += 1
            continue

        try:
            customer = Customer(
                tenant_id=current_user.tenant_id,
                created_by=current_user.id,
                name=name,
                **{k: v for k, v in customer_data.items() if hasattr(Customer, k)},
            )
            db.add(customer)
            created += 1
        except Exception as e:
            logger.warning(f"Import row {row_num} failed: {e}")
            errors.append(CustomerImportRowError(row=row_num, message=f"数据格式异常: {str(e)[:100]}"))
            skipped += 1

    await db.commit()

    return CustomerImportResponse(
        created=created,
        skipped=skipped,
        total=created + skipped,
        errors=errors,
    )


@router.get("/{customer_id}", response_model=CustomerDetailResponse)
async def get_customer(
    customer_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """客户详情（含联系人列表）"""
    uid = _parse_uuid(customer_id, "客户ID")
    result = await db.execute(
        select(Customer)
        .options(selectinload(Customer.contacts))
        .where(
            Customer.id == uid,
            Customer.tenant_id == current_user.tenant_id,
        )
    )
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    return CustomerDetailResponse(
        **_row_to_response(customer).model_dump(),
        contacts=[ContactResponse.model_validate(c) for c in customer.contacts],
    )


@router.put("/{customer_id}", response_model=CustomerResponse)
async def update_customer(
    customer_id: str,
    data: CustomerUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """更新客户信息"""
    uid = _parse_uuid(customer_id, "客户ID")
    result = await db.execute(
        select(Customer).where(
            Customer.id == uid,
            Customer.tenant_id == current_user.tenant_id,
        )
    )
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(customer, field, value)

    await db.commit()
    await db.refresh(customer)
    return _row_to_response(customer)


@router.delete("/{customer_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_customer(
    customer_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """删除客户（级联删除联系人）"""
    uid = _parse_uuid(customer_id, "客户ID")
    result = await db.execute(
        select(Customer).where(
            Customer.id == uid,
            Customer.tenant_id == current_user.tenant_id,
        )
    )
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    await db.delete(customer)
    await db.commit()
    return None


# ── 批量操作 ──

@router.put("/batch/status")
async def batch_update_status(
    data: BatchStatusRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """批量修改客户状态"""
    if not data.ids:
        raise HTTPException(status_code=400, detail="请选择至少一个客户")

    result = await db.execute(
        update(Customer)
        .where(
            Customer.id.in_(data.ids),
            Customer.tenant_id == current_user.tenant_id,
        )
        .values(status=data.status)
    )
    await db.commit()
    return {"updated_count": result.rowcount}


# ── 导出 Excel ──

@router.get("/export")
async def export_customers(
    status: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    icp_id: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """导出客户列表为 Excel 文件"""
    where_clause = _build_customer_filter(
        current_user.tenant_id, status, source, search, icp_id
    )

    q = (
        select(Customer)
        .where(where_clause)
        .options(selectinload(Customer.contacts))
        .order_by(Customer.created_at.desc())
        .limit(_MAX_EXPORT_ROWS)
    )
    rows = (await db.execute(q)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "客户列表"
    ws.append([
        "公司名称", "行业", "网站", "国家", "城市", "规模",
        "来源", "状态", "联系人数", "描述", "备注", "创建时间",
    ])

    for r in rows:
        ws.append([
            r.name, r.industry or "", r.website or "",
            r.country or "", r.city or "", r.company_size or "",
            r.source, r.status,
            len(r.contacts) if r.contacts else 0,
            r.description or "", r.notes or "",
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    date_str = datetime.now().strftime("%Y-%m-%d")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=customers_{date_str}.xlsx"
        },
    )


# ── 搜索核心逻辑（SSE + 后台共用）──

_CHANNELS_MAP = {
    "ai": AISearchChannel,
    "google": GoogleSearchChannel,
    "serper": SerperSearchChannel,
    "linkedin": LinkedInSearchChannel,
    "linkedin_people": LinkedInPeopleSearchChannel,
    "duckduckgo": DuckDuckGoSearchChannel,
}


async def _execute_search(
    tenant_id: _uuid.UUID,
    user_id: _uuid.UUID,
    icp_id: _uuid.UUID,
    icp_name: str,
    icp_data: dict,
    channels: List[str],
    region: str,
    *,
    on_progress,
):
    """执行 6 步客户搜索，通过 on_progress(event_dict) 报告进度。

    返回结果字典：{"saved_count", "enriched_count", "contact_search_count", "total_found"}
    """
    query = " ".join(filter(None, [
        icp_data.get("target_industry", ""),
        icp_data.get("product_category", ""),
    ])) or icp_name

    # Step 1: 各渠道搜索（并行）
    await on_progress({"type": "section", "section": "searching"})
    all_results: List[SearchResult] = []

    async def _search_channel(channel_name: str, channel_cls):
        """单个渠道搜索任务，返回 (channel_name, results, error_str)"""
        try:
            await on_progress({"type": "progress", "channel": channel_name, "message": f"正在 {channel_name} 搜索..."})
            channel = channel_cls()
            results = await channel.search(query, region, max_results=_MAX_SEARCH_RESULTS)
            await on_progress({"type": "progress", "channel": channel_name, "found": len(results), "message": f"{channel_name} 找到 {len(results)} 条"})
            return channel_name, results, None
        except Exception as e:
            logger.warning(f"Channel {channel_name} failed: {e}")
            await on_progress({"type": "progress", "channel": channel_name, "error": str(e), "message": f"{channel_name} 搜索失败: {e}"})
            return channel_name, [], str(e)

    channel_tasks = [
        _search_channel(name, cls)
        for name in channels
        if (cls := _CHANNELS_MAP.get(name))
    ]
    if channel_tasks:
        channel_results = await asyncio.gather(*channel_tasks)
        for _, results, _ in channel_results:
            all_results.extend(results)

    # Step 2: 聚合去重
    await on_progress({"type": "section", "section": "deduping"})
    aggregator = SearchAggregator()
    deduped = aggregator.aggregate(all_results)
    await on_progress({"type": "progress", "message": f"去重后 {len(deduped)} 条，开始 AI 结构化..."})

    # Step 3: AI 结构化并写入 DB（并行，最多 5 个并发提取）
    await on_progress({"type": "section", "section": "saving"})
    ai_service = get_ai_service()
    saved_count = 0
    saved_customers: list = []  # [(customer_id, website), ...]

    _extract_semaphore = asyncio.Semaphore(5)
    _saved_lock = asyncio.Lock()

    async def _extract_and_save(result: SearchResult) -> tuple:
        """单个结果的提取+保存任务。返回 (status, customer_id, website) 或 (status, None, None)"""
        nonlocal saved_count, saved_customers

        async with _extract_semaphore:
            try:
                if result.skip_extraction:
                    # AI 渠道已返回结构化数据，直接写入 DB
                    async with AsyncSessionLocal() as sess:
                        website_to_check = result.website
                        if website_to_check:
                            existing = await sess.execute(
                                select(func.count(Customer.id)).where(
                                    Customer.tenant_id == tenant_id,
                                    Customer.website == website_to_check,
                                )
                            )
                            if existing.scalar():
                                await on_progress({"type": "progress", "message": "已存在，跳过: " + result.company_name})
                                return ("skipped", None, None)

                        customer = Customer(
                            tenant_id=tenant_id, created_by=user_id,
                            name=result.company_name, industry=result.industry,
                            website=result.website, country=result.country,
                            city=result.city, company_size=result.company_size,
                            description=result.description,
                            source=(result.source_channel or "ai_search")[:48],
                            source_url=result.source_url, icp_id=icp_id,
                            status="new",
                            source_data={
                                "company_name": result.company_name,
                                "industry": result.industry,
                                "country": result.country,
                                "city": result.city,
                                "company_size": result.company_size,
                                "description": result.description,
                                "website": result.website,
                                "source": "ai_search",
                            },
                            ai_summary=result.description,
                        )
                        sess.add(customer)
                        await sess.flush()

                        for c_data in result.contacts or []:
                            if c_data.get("name"):
                                sess.add(Contact(
                                    customer_id=customer.id, tenant_id=tenant_id,
                                    name=c_data.get("name", ""), title=c_data.get("title"),
                                    email=c_data.get("email"), phone=c_data.get("phone"),
                                    linkedin_url=c_data.get("linkedin_url"),
                                    contact_type="ai_suggested",
                                    confidence=c_data.get("confidence", "inferred"),
                                ))

                        await sess.commit()
                        async with _saved_lock:
                            saved_count += 1
                            saved_customers.append((customer.id, result.website))
                        await on_progress({"type": "progress", "message": "已保存: " + result.company_name})
                        return ("saved", customer.id, result.website)

                else:
                    # 网页抓取渠道：抓取网页 + LLM 提取
                    page_content = None
                    page_fetched = False
                    if result.website and _is_safe_url(result.website):
                        try:
                            async with httpx.AsyncClient(timeout=_MAX_PAGE_FETCH_TIMEOUT, headers={
                                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
                            }) as client:
                                page_resp = await client.get(
                                    result.website if "://" in (result.website or "") else f"https://{result.website}"
                                )
                                if page_resp.status_code == 200:
                                    soup = BeautifulSoup(page_resp.text, "lxml")
                                    for tag in soup(["script", "style", "nav", "footer", "header"]):
                                        tag.decompose()
                                    page_content = soup.get_text(separator="\n", strip=True)[:10000]
                                    page_fetched = True
                        except Exception:
                            pass

                    if not page_fetched or not page_content:
                        await on_progress({"type": "progress", "message": "网页抓取失败，跳过: " + (result.company_name or result.website or "未知")})
                        return ("skipped", None, None)

                    extractor = CustomerExtractor(ai_service)
                    # 并行模式下不逐 token 输出流式文本（会交错混乱），仅静默提取
                    async for _ in extractor.extract(page_content, result.website or ""):
                        pass

                    output = extractor.get_output()
                    if output is None:
                        await on_progress({"type": "progress", "message": "非公司页面，跳过: " + (result.company_name or result.website or "未知")})
                        return ("skipped", None, None)

                    if not output.get("parse_error"):
                        company_name = output.get("company_name")
                        if not company_name:
                            await on_progress({"type": "progress", "message": "未提取到公司名，跳过: " + (result.website or "未知")})
                            return ("skipped", None, None)

                        async with AsyncSessionLocal() as sess:
                            website_to_check = output.get("website") or result.website
                            if website_to_check:
                                existing = await sess.execute(
                                    select(func.count(Customer.id)).where(
                                        Customer.tenant_id == tenant_id,
                                        Customer.website == website_to_check,
                                    )
                                )
                                if existing.scalar():
                                    await on_progress({"type": "progress", "message": "已存在，跳过: " + company_name})
                                    return ("skipped", None, None)

                            customer = Customer(
                                tenant_id=tenant_id, created_by=user_id,
                                name=company_name,
                                industry=output.get("industry") or result.industry,
                                website=output.get("website") or result.website,
                                country=output.get("country") or result.country,
                                city=output.get("city") or result.city,
                                company_size=output.get("company_size"),
                                description=output.get("description") or result.description,
                                source=(result.source_channel or "ai_search")[:48],
                                source_url=result.source_url, icp_id=icp_id,
                                status="new", source_data=output,
                                ai_summary=output.get("description"),
                            )
                            sess.add(customer)
                            await sess.flush()

                            for c_data in output.get("contacts") or []:
                                if c_data.get("name"):
                                    sess.add(Contact(
                                        customer_id=customer.id, tenant_id=tenant_id,
                                        name=c_data.get("name", ""), title=c_data.get("title"),
                                        email=c_data.get("email"), phone=c_data.get("phone"),
                                        linkedin_url=c_data.get("linkedin_url"),
                                        contact_type="scraped",
                                        confidence="high" if c_data.get("email") else "medium",
                                    ))

                            await sess.commit()
                            async with _saved_lock:
                                saved_count += 1
                                saved_customers.append((customer.id, output.get("website") or result.website))
                            await on_progress({"type": "progress", "message": "已保存: " + company_name})
                            return ("saved", customer.id, output.get("website") or result.website)

                    return ("skipped", None, None)

            except AIServiceError as e:
                logger.warning(f"AI extraction failed for {result.company_name}: {e}")
                return ("skipped", None, None)
            except Exception as e:
                logger.error(f"Unexpected error processing {result.company_name}: {e}")
                return ("skipped", None, None)

    await on_progress({"type": "progress", "message": f"并行提取 {len(deduped[:_MAX_SEARCH_RESULTS])} 条结果（最多 5 并发）..."})
    extract_tasks = [_extract_and_save(r) for r in deduped[:_MAX_SEARCH_RESULTS]]
    await asyncio.gather(*extract_tasks, return_exceptions=True)

    # Step 4: 定向联系人搜索（并行，最多 5 并发）
    await on_progress({"type": "section", "section": "contact_search"})
    contact_search_count = 0
    _contact_search_lock = asyncio.Lock()

    async def _contact_search_one(cid: str, website: str):
        """单家公司定向联系人搜索任务"""
        nonlocal contact_search_count
        try:
            async with AsyncSessionLocal() as sess:
                cust_result = await sess.execute(
                    select(Customer.name, Customer.website).where(Customer.id == cid)
                )
                cust_row = cust_result.first()
                if not cust_row:
                    return
                company_name = cust_row[0]
                domain = cust_row[1] or website

                searcher = ContactSearchService(timeout=15.0)
                found_contacts = await searcher.search_company_contacts(
                    company_name=company_name, domain=domain, region=region,
                )
                if found_contacts:
                    existing_result = await sess.execute(
                        select(Contact.name).where(Contact.customer_id == cid)
                    )
                    existing_names = {r[0].lower().strip() for r in existing_result.all() if r[0]}

                    added = 0
                    for c_data in found_contacts:
                        c_name = (c_data.get("name") or "").strip()
                        if c_name and c_name.lower() in existing_names:
                            continue
                        if not c_data.get("email") and not c_name:
                            continue

                        sess.add(Contact(
                            customer_id=cid, tenant_id=tenant_id,
                            name=c_name or (c_data.get("email", "").split("@")[0] if c_data.get("email") else "Unknown"),
                            title=c_data.get("title"), email=c_data.get("email"),
                            phone=c_data.get("phone"), linkedin_url=c_data.get("linkedin_url"),
                            contact_type="ai_suggested", confidence="medium",
                        ))
                        if c_name:
                            existing_names.add(c_name.lower())
                        added += 1

                    if added > 0:
                        await sess.commit()
                        async with _contact_search_lock:
                            contact_search_count += 1
                        await on_progress({"type": "progress", "message": f"从定向搜索为 {company_name} 找到 {added} 个联系人"})
        except Exception as e:
            logger.warning(f"Contact search failed for customer {cid}: {e}")

    if saved_customers:
        await on_progress({"type": "progress", "message": f"正在对 {len(saved_customers)} 家公司进行定向联系人搜索（并行）..."})
        _cs_sem = asyncio.Semaphore(5)
        async def _cs_limited(cid, website):
            async with _cs_sem:
                await _contact_search_one(cid, website)
        await asyncio.gather(*[_cs_limited(cid, ws) for cid, ws in saved_customers], return_exceptions=True)

    await on_progress({"type": "progress", "message": f"定向搜索完成，为 {contact_search_count} 家公司找到联系人"})

    # Step 5: 批量爬虫补全联系人（并行，最多 3 并发——爬虫较重）
    await on_progress({"type": "section", "section": "enriching"})
    enriched_count = 0
    _enrich_lock = asyncio.Lock()

    async def _scrape_one(cid: str, website: str):
        """单个网站联系人抓取任务：网站抓取 + Hunter.io（如已配置）"""
        nonlocal enriched_count
        if not website:
            return

        all_scraped_contacts: list = []

        # 1. 网站抓取
        try:
            scraper = ContactScraper(timeout=30.0)
            contacts = await scraper.scrape(website)
            if contacts:
                all_scraped_contacts.extend(contacts)
        except Exception as e:
            logger.warning(f"Contact scraping failed for {website}: {e}")

        # 2. Hunter.io 域名搜索（如果配置了 API Key）
        try:
            hunter = HunterService()
            if hunter.is_available:
                domain = website.replace("https://", "").replace("http://", "").rstrip("/").split("/")[0]
                hunter_contacts = await hunter.domain_search(domain)
                if hunter_contacts:
                    # 按 name+email 去重（与已有结果合并）
                    seen_keys = {
                        (c.get("name", "").lower(), (c.get("email") or "").lower())
                        for c in all_scraped_contacts
                    }
                    for hc in hunter_contacts:
                        key = ((hc.get("name") or "").lower(), (hc.get("email") or "").lower())
                        if key[1] and key not in seen_keys:  # 只添加有邮箱的 Hunter 结果
                            seen_keys.add(key)
                            all_scraped_contacts.append(hc)
                    await on_progress({"type": "progress", "message": f"Hunter 为 {domain} 找到 {len(hunter_contacts)} 个邮箱"})
        except Exception as e:
            logger.warning(f"Hunter search failed for {website}: {e}")

        # 3. 保存到数据库
        if all_scraped_contacts:
            valid_contacts = [
                c for c in all_scraped_contacts
                if c.get("email") or (c.get("name") and c.get("title"))
            ]
            if valid_contacts:
                async with AsyncSessionLocal() as sess:
                    existing_result = await sess.execute(
                        select(Contact.name).where(Contact.customer_id == cid)
                    )
                    existing_names = {r[0].lower().strip() for r in existing_result.all() if r[0]}

                    added = 0
                    for c_data in valid_contacts:
                        c_name = (c_data.get("name") or "").strip()
                        if c_name and c_name.lower() in existing_names:
                            continue
                        if not c_data.get("email") and not c_name:
                            continue

                        sess.add(Contact(
                            customer_id=cid, tenant_id=tenant_id,
                            name=c_name or c_data.get("email", "").split("@")[0],
                            title=c_data.get("title"), email=c_data.get("email"),
                            phone=c_data.get("phone"), linkedin_url=c_data.get("linkedin_url"),
                            contact_type="scraped",
                            confidence="high" if c_data.get("email") else "medium",
                        ))
                        if c_name:
                            existing_names.add(c_name.lower())
                        added += 1

                    if added > 0:
                        await sess.commit()
                        async with _enrich_lock:
                            enriched_count += 1
                        await on_progress({"type": "progress", "message": f"从 {website} 找到 {added} 个联系人"})

    if saved_customers:
        await on_progress({"type": "progress", "message": f"正在从 {len(saved_customers)} 个公司网站抓取联系人信息（并行）..."})
        _scrape_sem = asyncio.Semaphore(3)
        async def _scrape_limited(cid, website):
            async with _scrape_sem:
                await _scrape_one(cid, website)
        await asyncio.gather(*[_scrape_limited(cid, ws) for cid, ws in saved_customers], return_exceptions=True)

    result = {
        "saved_count": saved_count,
        "enriched_count": enriched_count,
        "contact_search_count": contact_search_count,
        "total_found": len(deduped),
    }
    await on_progress({"type": "complete", **result})
    return result


# ── 搜索 SSE ──

@router.post("/search")
async def search_customers(
    req: CustomerSearchRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """SSE 客户搜索：基于画像关键词，通过 Google/LinkedIn 搜索 → 聚合去重 → AI 结构化"""

    icp_result = await db.execute(
        select(Icp).where(Icp.id == req.icp_id, Icp.tenant_id == current_user.tenant_id)
    )
    icp = icp_result.scalar_one_or_none()
    if not icp:
        raise HTTPException(status_code=404, detail="画像不存在")

    icp_data = icp.input_data or {}
    region = req.region or icp_data.get("target_region", "")
    _tenant_id = current_user.tenant_id
    _user_id = current_user.id
    _icp_id = icp.id
    _icp_name = icp.name

    async def event_stream():
        queue: asyncio.Queue = asyncio.Queue()

        async def sse_callback(event: dict):
            await queue.put(event)

        # 启动搜索任务（后台 asyncio task，通过 queue 传递进度）
        search_task = asyncio.create_task(
            _execute_search(
                tenant_id=_tenant_id, user_id=_user_id, icp_id=_icp_id,
                icp_name=_icp_name, icp_data=icp_data,
                channels=req.channels, region=region,
                on_progress=sse_callback,
            )
        )

        try:
            while True:
                try:
                    event = await asyncio.wait_for(queue.get(), timeout=0.1)
                    yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                    if event.get("type") in ("complete", "error"):
                        break
                except asyncio.TimeoutError:
                    if search_task.done():
                        # 任务已完成但可能没有 complete 事件（异常情况）
                        if search_task.exception():
                            yield f"data: {json.dumps({'type': 'error', 'message': f'搜索失败: {search_task.exception()}'}, ensure_ascii=False)}\n\n"
                        break
        except Exception as e:
            logger.error(f"Search SSE fatal error: {e}")
            yield f"data: {json.dumps({'type': 'error', 'message': f'搜索失败: {e}'}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ── 后台搜索 ──

from app.services.search.task_manager import search_task_manager


@router.post("/search/background")
async def search_customers_background(
    req: CustomerSearchRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """后台客户搜索：立即返回 task_id，搜索在后台执行"""

    icp_result = await db.execute(
        select(Icp).where(Icp.id == req.icp_id, Icp.tenant_id == current_user.tenant_id)
    )
    icp = icp_result.scalar_one_or_none()
    if not icp:
        raise HTTPException(status_code=404, detail="画像不存在")

    icp_data = icp.input_data or {}
    region = req.region or icp_data.get("target_region", "")
    tenant_id_str = str(current_user.tenant_id)

    task_id = await search_task_manager.create(
        tenant_id=tenant_id_str,
        icp_name=icp.name,
        channels=req.channels,
        region=region,
    )

    async def background_callback(event: dict):
        """将搜索进度更新到任务管理器"""
        task = await search_task_manager.get(task_id)
        if task and task["status"] == "pending":
            await search_task_manager.update(task_id, status="running")

        if event.get("type") == "section":
            await search_task_manager.update(task_id, current_section=event.get("section"))
        elif event.get("type") == "progress":
            await search_task_manager.update(task_id, progress_message=event.get("message", ""))
        elif event.get("type") == "complete":
            await search_task_manager.update(
                task_id,
                status="completed",
                progress_message="搜索完成",
                result={
                    "saved_count": event.get("saved_count", 0),
                    "enriched_count": event.get("enriched_count", 0),
                    "contact_search_count": event.get("contact_search_count", 0),
                    "total_found": event.get("total_found", 0),
                },
            )
        elif event.get("type") == "error":
            await search_task_manager.update(
                task_id, status="failed",
                error=event.get("message", "未知错误"),
                progress_message="搜索失败",
            )

    async def run_background():
        try:
            await _execute_search(
                tenant_id=current_user.tenant_id,
                user_id=current_user.id,
                icp_id=icp.id,
                icp_name=icp.name,
                icp_data=icp_data,
                channels=req.channels,
                region=region,
                on_progress=background_callback,
            )
        except Exception as e:
            logger.error(f"Background search task {task_id} failed: {e}")
            await search_task_manager.update(
                task_id, status="failed",
                error=str(e),
                progress_message=f"搜索失败: {e}",
            )

    asyncio.create_task(run_background())
    return {"task_id": task_id}


@router.get("/search/tasks")
async def list_search_tasks(
    current_user: User = Depends(get_current_user),
):
    """获取当前租户的后台搜索任务列表"""
    tasks = await search_task_manager.list_by_tenant(str(current_user.tenant_id))
    return {"tasks": tasks}


# ── 信息补全（Phase 5） ──

@router.post("/{customer_id}/enrich")
async def enrich_customer(
    customer_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """SSE 端点：补全客户信息（联系人与字段）"""
    uid = _parse_uuid(customer_id, "客户ID")
    # 验证归属
    result = await db.execute(
        select(Customer).where(
            Customer.id == uid,
            Customer.tenant_id == current_user.tenant_id,
        )
    )
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    async def event_stream():
        enricher = CustomerEnricher()
        async for event in enricher.enrich(
            customer_id=str(uid),
            tenant_id=str(current_user.tenant_id),
            db=db,
        ):
            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.post("/batch/enrich")
async def batch_enrich_customers(
    data: BatchStatusRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """批量触发客户信息补全"""
    if not data.ids:
        raise HTTPException(status_code=400, detail="请选择至少一个客户")

    # 验证所有 ID 归属
    result = await db.execute(
        select(func.count(Customer.id)).where(
            Customer.id.in_(data.ids),
            Customer.tenant_id == current_user.tenant_id,
        )
    )
    valid_count = result.scalar() or 0

    # 标记为 pending
    await db.execute(
        update(Customer)
        .where(
            Customer.id.in_(data.ids),
            Customer.tenant_id == current_user.tenant_id,
        )
        .values(enrichment_status="pending")
    )
    await db.commit()

    return {
        "queued": valid_count,
        "total": len(data.ids),
        "message": f"已标记 {valid_count} 个客户待补全，请逐个触发或稍后查看",
    }


# ── 联系人子路由 ──

@router.get("/{customer_id}/contacts", response_model=List[ContactResponse])
async def list_contacts(
    customer_id: str,
    _dummy: None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """获取客户的联系人列表"""
    uid = _parse_uuid(customer_id, "客户ID")
    # 先验证客户归属
    cust_result = await db.execute(
        select(Customer.id).where(
            Customer.id == uid,
            Customer.tenant_id == current_user.tenant_id,
        )
    )
    if not cust_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="客户不存在")

    result = await db.execute(
        select(Contact)
        .where(Contact.customer_id == uid)
        .order_by(Contact.is_primary.desc(), Contact.created_at.asc())
    )
    return [ContactResponse.model_validate(c) for c in result.scalars().all()]


@router.post("/{customer_id}/contacts", response_model=ContactResponse, status_code=status.HTTP_201_CREATED)
async def create_contact(
    customer_id: str,
    data: ContactCreateRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """为客户添加联系人"""
    uid = _parse_uuid(customer_id, "客户ID")
    cust_result = await db.execute(
        select(Customer).where(
            Customer.id == uid,
            Customer.tenant_id == current_user.tenant_id,
        )
    )
    if not cust_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="客户不存在")

    contact = Contact(
        customer_id=uid,
        tenant_id=current_user.tenant_id,
        **data.model_dump(),
    )
    db.add(contact)
    await db.commit()
    await db.refresh(contact)
    return contact


@router.put("/{customer_id}/contacts/{contact_id}", response_model=ContactResponse)
async def update_contact(
    customer_id: str,
    contact_id: str,
    data: ContactUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """更新联系人"""
    uid = _parse_uuid(customer_id, "客户ID")
    cid = _parse_uuid(contact_id, "联系人ID")
    result = await db.execute(
        select(Contact).where(
            Contact.id == cid,
            Contact.customer_id == uid,
            Contact.tenant_id == current_user.tenant_id,
        )
    )
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="联系人不存在")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(contact, field, value)

    await db.commit()
    await db.refresh(contact)
    return contact


@router.delete("/{customer_id}/contacts/{contact_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_contact(
    customer_id: str,
    contact_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """删除联系人"""
    uid = _parse_uuid(customer_id, "客户ID")
    cid = _parse_uuid(contact_id, "联系人ID")
    result = await db.execute(
        select(Contact).where(
            Contact.id == cid,
            Contact.customer_id == uid,
            Contact.tenant_id == current_user.tenant_id,
        )
    )
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="联系人不存在")

    await db.delete(contact)
    await db.commit()


# ── AI 搜联系人（v1.4） ──

@router.post("/{customer_id}/ai-search-contacts")
async def ai_search_contacts(
    customer_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """SSE 端点：AI 搜索联系人 — 多数据源搜索 + 抓取 + 推断"""
    uid = _parse_uuid(customer_id, "客户ID")
    result = await db.execute(
        select(Customer).where(
            Customer.id == uid,
            Customer.tenant_id == current_user.tenant_id,
        )
    )
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    _tenant_id = current_user.tenant_id
    _customer_id = customer.id
    company_name = customer.name
    website = customer.website
    domain = None
    if website:
        parsed = urlparse(website if "://" in website else f"https://{website}")
        domain = parsed.netloc.replace("www.", "") if parsed.netloc else None

    # 已有联系人用于去重
    existing_contacts_result = await db.execute(
        select(Contact.name, Contact.email).where(Contact.customer_id == uid)
    )
    existing_rows = existing_contacts_result.all()
    existing_names = {r[0].lower().strip() for r in existing_rows if r[0]}
    existing_emails = {r[1].lower() for r in existing_rows if r[1]}

    async def event_stream():
        linkedin_found = 0
        contact_search_found = 0
        scraped_found = 0

        try:
            # Step 1: LinkedIn 人物搜索
            yield f"data: {json.dumps({'type': 'section', 'section': 'linkedin_people'}, ensure_ascii=False)}\n\n"
            yield f"data: {json.dumps({'type': 'progress', 'message': f'正在 LinkedIn 搜索 {company_name} 的相关人员...'}, ensure_ascii=False)}\n\n"

            try:
                linkedin_channel = LinkedInPeopleSearchChannel(timeout=15.0)
                linkedin_results = await linkedin_channel.search(
                    query=company_name,
                    region=customer.country or "",
                    max_results=10,
                )
                new_contacts_added = 0
                for r in linkedin_results:
                    for c_data in r.contacts or []:
                        c_name = (c_data.get("name") or "").strip()
                        if not c_name or c_name.lower() in existing_names:
                            continue
                        async with AsyncSessionLocal() as sess:
                            contact = Contact(
                                customer_id=_customer_id,
                                tenant_id=_tenant_id,
                                name=c_name,
                                title=c_data.get("title"),
                                email=c_data.get("email"),
                                phone=c_data.get("phone"),
                                linkedin_url=c_data.get("linkedin_url") or r.source_url,
                                contact_type="ai_suggested",
                                confidence="medium",
                            )
                            sess.add(contact)
                            await sess.commit()
                        existing_names.add(c_name.lower())
                        new_contacts_added += 1
                linkedin_found = new_contacts_added
                yield f"data: {json.dumps({'type': 'progress', 'message': f'LinkedIn 找到 {linkedin_found} 个新联系人'}, ensure_ascii=False)}\n\n"
            except Exception as e:
                logger.warning(f"LinkedIn people search failed: {e}")
                yield f"data: {json.dumps({'type': 'progress', 'message': f'LinkedIn 搜索失败: {e}'}, ensure_ascii=False)}\n\n"

            # Step 2: 定向联系人搜索
            yield f"data: {json.dumps({'type': 'section', 'section': 'contact_search'}, ensure_ascii=False)}\n\n"
            yield f"data: {json.dumps({'type': 'progress', 'message': f'正在执行定向联系人搜索...'}, ensure_ascii=False)}\n\n"

            try:
                contact_searcher = ContactSearchService(timeout=15.0)
                found_contacts = await contact_searcher.search_company_contacts(
                    company_name=company_name,
                    domain=website or domain,
                    region=customer.country or "",
                )
                added = 0
                for c_data in found_contacts:
                    c_name = (c_data.get("name") or "").strip()
                    if c_name and c_name.lower() in existing_names:
                        continue
                    if not c_data.get("email") and not c_name:
                        continue
                    async with AsyncSessionLocal() as sess:
                        contact = Contact(
                            customer_id=_customer_id,
                            tenant_id=_tenant_id,
                            name=c_name or (c_data.get("email", "").split("@")[0] if c_data.get("email") else "Unknown"),
                            title=c_data.get("title"),
                            email=c_data.get("email"),
                            phone=c_data.get("phone"),
                            linkedin_url=c_data.get("linkedin_url"),
                            contact_type="ai_suggested",
                            confidence="medium" if c_data.get("email") else "low",
                        )
                        sess.add(contact)
                        await sess.commit()
                    if c_name:
                        existing_names.add(c_name.lower())
                    added += 1
                contact_search_found = added
                yield f"data: {json.dumps({'type': 'progress', 'message': f'定向搜索找到 {contact_search_found} 个联系人'}, ensure_ascii=False)}\n\n"
            except Exception as e:
                logger.warning(f"Contact search failed: {e}")
                yield f"data: {json.dumps({'type': 'progress', 'message': f'定向搜索失败: {e}'}, ensure_ascii=False)}\n\n"

            # Step 3: 网站抓取联系人
            if website:
                yield f"data: {json.dumps({'type': 'section', 'section': 'scraping'}, ensure_ascii=False)}\n\n"
                yield f"data: {json.dumps({'type': 'progress', 'message': f'正在从 {website} 抓取联系人信息...'}, ensure_ascii=False)}\n\n"

                try:
                    scraper = ContactScraper(timeout=30.0)
                    scraped_contacts = await scraper.scrape(website)
                    valid_contacts = [
                        c for c in scraped_contacts
                        if c.get("email") or (c.get("name") and c.get("title"))
                    ]
                    added = 0
                    for c_data in valid_contacts:
                        c_name = (c_data.get("name") or "").strip()
                        if c_name and c_name.lower() in existing_names:
                            continue
                        if not c_data.get("email") and not c_name:
                            continue
                        async with AsyncSessionLocal() as sess:
                            contact = Contact(
                                customer_id=_customer_id,
                                tenant_id=_tenant_id,
                                name=c_name or (c_data.get("email", "").split("@")[0] if c_data.get("email") else "Unknown"),
                                title=c_data.get("title"),
                                email=c_data.get("email"),
                                phone=c_data.get("phone"),
                                linkedin_url=c_data.get("linkedin_url"),
                                contact_type="scraped",
                                confidence="high" if c_data.get("email") else "medium",
                            )
                            sess.add(contact)
                            await sess.commit()
                        if c_name:
                            existing_names.add(c_name.lower())
                        if c_data.get("email"):
                            existing_emails.add(c_data["email"].lower())
                        added += 1
                    scraped_found = added
                    yield f"data: {json.dumps({'type': 'progress', 'message': f'网站抓取找到 {scraped_found} 个联系人'}, ensure_ascii=False)}\n\n"
                except Exception as e:
                    logger.warning(f"Contact scraping failed: {e}")
                    yield f"data: {json.dumps({'type': 'progress', 'message': f'网站抓取失败: {e}'}, ensure_ascii=False)}\n\n"

            total = linkedin_found + contact_search_found + scraped_found
            yield f"data: {json.dumps({'type': 'complete', 'linkedin_found': linkedin_found, 'contact_search_found': contact_search_found, 'scraped_found': scraped_found, 'total': total}, ensure_ascii=False)}\n\n"

        except Exception as e:
            logger.error(f"AI search contacts fatal error: {e}")
            yield f"data: {json.dumps({'type': 'error', 'message': f'搜索失败: {e}'}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )
    return None
